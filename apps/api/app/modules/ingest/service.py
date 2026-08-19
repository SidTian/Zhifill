from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from docx import Document
from docx.document import Document as DocxDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader

from aff_contracts import DocumentBundle, IngestRequest
from aff_contracts.ingest import ChunkHint, TableBlock
from app.core.errors import NotImplementedModule
from app.modules.ingest.port import IngestPort


def iter_docx_blocks(
    document: DocxDocument,
) -> Iterator[Paragraph | Table]:
    """
    Iterate through Word paragraphs and tables
    in their original document order.
    """
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(
                child,
                document,
            )

        elif isinstance(child, CT_Tbl):
            yield Table(
                child,
                document,
            )


class IngestService(IngestPort):
    def ingest(self, request: IngestRequest) -> DocumentBundle:
        path = Path(request.file.path)
        suffix = path.suffix.lower()

        tables: list[TableBlock] | None = None
        chunks_hint: list[ChunkHint] | None = None
        warnings: list[str] = []

        metadata = {
            "filename": request.file.filename,
            "mime": request.file.mime,
            "sha256": request.file.sha256,
            "size": request.file.size,
        }

        # TXT / Markdown
        if suffix in {".txt", ".md"}:
            text = path.read_text(
                encoding="utf-8",
            ).strip()

        # Word
        elif suffix == ".docx":
            document = Document(path)

            tables = []
            text_parts: list[str] = []

            table_index = 0

            for block in iter_docx_blocks(document):
                # 普通段落
                if isinstance(block, Paragraph):
                    paragraph_text = block.text.strip()

                    if paragraph_text:
                        text_parts.append(
                            paragraph_text
                        )

                # Word 表格
                elif isinstance(block, Table):
                    table_index += 1

                    table_rows: list[list[str]] = []

                    for row in block.rows:
                        values = [
                            cell.text.strip()
                            for cell in row.cells
                        ]

                        if any(values):
                            table_rows.append(values)

                    if not table_rows:
                        continue

                    table_name = (
                        f"Table {table_index}"
                    )

                    tables.append(
                        TableBlock(
                            name=table_name,
                            headers=table_rows[0],
                            rows=table_rows[1:],
                        )
                    )

                    rows_as_text = [
                        "\t".join(row)
                        for row in table_rows
                    ]

                    table_text = (
                        f"[Table: {table_name}]\n"
                        + "\n".join(rows_as_text)
                    )

                    text_parts.append(
                        table_text
                    )

            text = "\n".join(
                text_parts
            ).strip()

        # Excel
        elif suffix == ".xlsx":
            workbook = load_workbook(
                path,
                data_only=False,
                read_only=True,
            )

            metadata["sheet_names"] = [
                sheet.title
                for sheet in workbook.worksheets
            ]

            sheet_texts = []
            tables = []
            chunks_hint = []

            try:
                for sheet in workbook.worksheets:
                    sheet_rows: list[list[str]] = []

                    for row in sheet.iter_rows(
                        values_only=True,
                    ):
                        values = [
                            str(value).strip()
                            if value is not None
                            else ""
                            for value in row
                        ]

                        if any(values):
                            sheet_rows.append(values)

                    if not sheet_rows:
                        warnings.append(
                            f"Excel sheet "
                            f"'{sheet.title}' is empty."
                        )
                        continue

                    rows_as_text = [
                        "\t".join(row)
                        for row in sheet_rows
                    ]

                    sheet_body = "\n".join(
                        rows_as_text
                    )

                    sheet_texts.append(
                        f"[Sheet: {sheet.title}]\n"
                        + sheet_body
                    )

                    tables.append(
                        TableBlock(
                            name=sheet.title,
                            headers=sheet_rows[0],
                            rows=sheet_rows[1:],
                        )
                    )

                    chunks_hint.append(
                        ChunkHint(
                            text=sheet_body,
                            sheet=sheet.title,
                        )
                    )

                text = "\n\n".join(
                    sheet_texts
                ).strip()

            finally:
                workbook.close()

        # PDF
        elif suffix == ".pdf":
            reader = PdfReader(path)

            metadata["page_count"] = len(
                reader.pages
            )

            pages = []
            chunks_hint = []

            for page_number, page in enumerate(
                reader.pages,
                start=1,
            ):
                page_text = page.extract_text()

                if page_text:
                    page_text = page_text.strip()

                if not page_text:
                    warnings.append(
                        f"PDF page {page_number} "
                        f"contains no extractable text."
                    )
                    continue

                pages.append(page_text)

                chunks_hint.append(
                    ChunkHint(
                        text=page_text,
                        page=page_number,
                    )
                )

            text = "\n\n".join(
                pages
            ).strip()

        # Unsupported
        else:
            raise NotImplementedModule(
                "ingest",
                f"unsupported file type: "
                f"{path.suffix}",
            )

        return DocumentBundle(
            doc_id=request.doc_id,
            title=path.stem,
            media_type=request.file.mime,
            text=text,
            chunks_hint=chunks_hint,
            tables=tables,
            metadata=metadata,
            warnings=warnings,
        )


def get_ingest_service() -> IngestPort:
    return IngestService()
